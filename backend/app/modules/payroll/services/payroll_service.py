"""
급여 서비스
──────────────────────────────────────────────────
- 관리자: 전직원 급여 조회/수정/엑셀 다운로드
- 직원:  본인 급여 명세서 조회

급여 계산 기준 (한국 노동법):
  주간급여   = wage × day_hours
  야간급여   = wage × night_hours × 1.5  (야간가산 포함)
  주휴수당   = wage × weekly_allowance_hours
  연차수당   = wage × annual_leave_hours
  공휴일수당 = wage × holiday_hours × 1.5
  근로자의날 = wage × labor_day_hours × 1.5

공제 (연도별 InsuranceRate 기준):
  건강보험   = gross × health_insurance_rate / 100
  요양보험   = 건강보험료 × long_term_care_rate / 100
  고용보험   = gross × employment_insurance_rate / 100
  국민연금   = gross × national_pension_rate / 100
"""

from __future__ import annotations

import io
from datetime import date, timedelta
from decimal import ROUND_DOWN, Decimal
from typing import List, Optional, Union

import openpyxl
from sqlalchemy.orm import Session

from app.modules.admin.models import InsuranceRate
from app.modules.auth.models import User
from app.modules.auth.services import decrypt_ssn
from app.modules.payroll.models import Payroll, PayrollPayDate
from app.modules.payroll.schemas import (
    PayrollAdminUpdateInput,
    PayrollPayResponse,
    PayrollResponse,
)
from app.modules.workstatus.models import AttendanceEvent, EventType
from app.utils.permission_utils import is_admin


class PayrollService:

    @staticmethod
    def get_payrolls(
        *,
        db: Session,
        user: User,
        year: int,
        month: Optional[int] = None,
    ) -> Union[List[PayrollResponse], PayrollPayResponse]:
        """급여 조회 (관리자: 전체, 직원: 본인)"""

        if is_admin(user):
            query = (
                db.query(Payroll)
                .join(User, Payroll.user_id == User.id)
                .filter(Payroll.year == year)
            )
            if month is not None:
                query = query.filter(Payroll.month == month)

            payrolls = query.order_by(User.name).all()
            return [PayrollService._to_admin_response(p, db) for p in payrolls]

        # 직원 본인
        payroll = (
            db.query(Payroll)
            .filter(
                Payroll.user_id == user.id,
                Payroll.year == year,
                Payroll.month == month,
            )
            .first()
        )
        if not payroll:
            return PayrollPayResponse()

        return PayrollService._to_user_pay_response(payroll, db)

    @staticmethod
    def update_payroll(
        *,
        db: Session,
        payroll_id: int,
        data: PayrollAdminUpdateInput,
    ) -> PayrollResponse:
        """관리자 급여 항목 수정 (총급여·총공제·실수령액 제외)"""
        payroll = db.get(Payroll, payroll_id)
        if not payroll:
            from fastapi import HTTPException
            raise HTTPException(status_code=404, detail="급여 기록을 찾을 수 없습니다.")

        update_fields = data.model_dump(exclude_none=True)
        for field, value in update_fields.items():
            setattr(payroll, field, value)

        db.commit()
        db.refresh(payroll)
        return PayrollService._to_admin_response(payroll, db)

    # ──────────────────────────────────────────────────────
    # 내부: 관리자 응답 생성
    # ──────────────────────────────────────────────────────
    @staticmethod
    def _to_admin_response(payroll: Payroll, db: Session) -> PayrollResponse:
        user = payroll.user

        w = payroll.wage

        day_wage = int(w * float(payroll.day_hours))
        night_wage = int(w * float(payroll.night_hours) * 1.5)
        weekly_allowance_pay = int(w * float(payroll.weekly_allowance_hours))
        annual_leave_pay = int(w * float(payroll.annual_leave_hours))
        holiday_pay = int(w * float(payroll.holiday_hours) * 1.5)
        labor_day_pay = int(w * float(payroll.labor_day_hours) * 1.5)

        gross_pay = (
            day_wage
            + night_wage
            + weekly_allowance_pay
            + annual_leave_pay
            + holiday_pay
            + labor_day_pay
        )

        total_work_hours = float(payroll.day_hours) + float(payroll.night_hours)

        # 보험료 계산
        rate_year = _insurance_rate_year(payroll.year, payroll.month)
        rate = _get_insurance_rate(db, rate_year)
        health, care, employment, pension = _calc_insurance(
            gross_pay,
            payroll.insurance_health,
            payroll.insurance_care,
            payroll.insurance_employment,
            payroll.insurance_pension,
            rate,
        )
        total_deduction = health + care + employment + pension

        # 근무일수 집계
        total_work_days = _count_work_days(db, payroll.user_id, payroll.year, payroll.month)

        # SSN 복호화
        raw_ssn = None
        if user.ssn:
            try:
                raw_ssn = decrypt_ssn(user.ssn)
            except Exception:
                raw_ssn = user.ssn

        return PayrollResponse(
            payroll_id=payroll.id,
            user_id=user.id,
            # 인적 정보
            name=user.name,
            position=user.position.value if user.position else None,
            wage=payroll.wage,
            rrn=raw_ssn,
            join_date=user.hire_date,
            resign_date=user.retire_date,
            last_work_day=payroll.last_work_day,
            bank_name=user.bank_name,
            bank_account=user.account_number,
            email=user.email,
            # 근무 요약
            total_work_days=total_work_days,
            total_work_hours=round(total_work_hours, 2),
            avg_daily_hours=(
                round(total_work_hours / total_work_days, 2)
                if total_work_days > 0 else None
            ),
            # 시간 항목
            day_hours=float(payroll.day_hours),
            night_hours=float(payroll.night_hours),
            weekly_allowance_hours=float(payroll.weekly_allowance_hours),
            annual_leave_hours=float(payroll.annual_leave_hours),
            holiday_hours=float(payroll.holiday_hours),
            labor_day_hours=float(payroll.labor_day_hours),
            # 급여 항목
            day_wage=day_wage,
            night_wage=night_wage,
            weekly_allowance_pay=weekly_allowance_pay,
            annual_leave_pay=annual_leave_pay,
            holiday_pay=holiday_pay,
            labor_day_pay=labor_day_pay,
            gross_pay=gross_pay,
            # 공제
            insurance_health=int(health),
            insurance_care=int(care),
            insurance_employment=int(employment),
            insurance_pension=int(pension),
            total_deduction=int(total_deduction),
            net_pay=gross_pay - int(total_deduction),
        )

    # ──────────────────────────────────────────────────────
    # 내부: 직원 응답 생성
    # ──────────────────────────────────────────────────────
    @staticmethod
    def _to_user_pay_response(payroll: Payroll, db: Session) -> PayrollPayResponse:
        user = payroll.user
        pay_date = _get_pay_date(db, payroll.year, payroll.month)

        w = payroll.wage

        day_pay = int(w * float(payroll.day_hours))
        night_pay = int(w * float(payroll.night_hours) * 1.5)
        weekly_allowance_pay = int(w * float(payroll.weekly_allowance_hours))
        annual_leave_pay = int(w * float(payroll.annual_leave_hours))
        holiday_pay = int(w * float(payroll.holiday_hours) * 1.5)
        labor_day_pay = int(w * float(payroll.labor_day_hours) * 1.5)

        gross_pay = (
            day_pay
            + night_pay
            + weekly_allowance_pay
            + annual_leave_pay
            + holiday_pay
            + labor_day_pay
        )

        rate_year = _insurance_rate_year(payroll.year, payroll.month)
        rate = _get_insurance_rate(db, rate_year)
        health, care, employment, pension = _calc_insurance(
            gross_pay,
            payroll.insurance_health,
            payroll.insurance_care,
            payroll.insurance_employment,
            payroll.insurance_pension,
            rate,
        )
        total_deduction = health + care + employment + pension

        total_work_hours = float(payroll.day_hours) + float(payroll.night_hours)
        total_work_days = _count_work_days(db, payroll.user_id, payroll.year, payroll.month)

        return PayrollPayResponse(
            name=user.name,
            position=user.position.value if user.position else None,
            birth_date=user.birth_date,
            pay_date=pay_date,
            wage=payroll.wage,
            # 시간 항목
            total_work_days=total_work_days,
            total_work_hours=round(total_work_hours, 2),
            avg_daily_hours=(
                round(total_work_hours / total_work_days, 2)
                if total_work_days > 0 else None
            ),
            day_hours=float(payroll.day_hours),
            night_hours=float(payroll.night_hours),
            weekly_allowance_hours=float(payroll.weekly_allowance_hours),
            annual_leave_hours=float(payroll.annual_leave_hours),
            holiday_hours=float(payroll.holiday_hours),
            labor_day_hours=float(payroll.labor_day_hours),
            # 급여 항목
            day_wage=day_pay,
            night_wage=night_pay,
            weekly_allowance_pay=weekly_allowance_pay,
            annual_leave_pay=annual_leave_pay,
            holiday_pay=holiday_pay,
            labor_day_pay=labor_day_pay,
            gross_pay=gross_pay,
            # 공제
            insurance_health=int(health),
            insurance_care=int(care),
            insurance_employment=int(employment),
            insurance_pension=int(pension),
            total_deduction=int(total_deduction),
            net_pay=gross_pay - int(total_deduction),
        )

    # ──────────────────────────────────────────────────────
    # 엑셀 내보내기
    # ──────────────────────────────────────────────────────
    @staticmethod
    def export_to_excel(
        payrolls: List[PayrollResponse],
        year: int,
        month: int,
    ) -> io.BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{year}년{month}월 급여"

        headers = [
            "이름", "직급", "주민등록번호", "시급",
            "입사일", "퇴사일", "마지막근무일",
            "근무일수", "총근무시간", "일평균시간",
            "주간시간", "야간시간", "주휴시간", "연차시간", "공휴일시간", "근로자의날시간",
            "주간급여", "야간급여", "주휴수당", "연차수당", "공휴일수당", "근로자의날수당",
            "급여총액",
            "건강보험", "요양보험", "고용보험", "국민연금",
            "공제계", "실수령액",
        ]
        ws.append(headers)

        # 헤더 스타일
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_fill = PatternFill(start_color="1A0F3C", end_color="1A0F3C", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for p in payrolls:
            ws.append([
                p.name, p.position, p.rrn or "", p.wage,
                str(p.join_date) if p.join_date else "",
                str(p.resign_date) if p.resign_date else "",
                str(p.last_work_day) if p.last_work_day else "",
                p.total_work_days or 0,
                p.total_work_hours or 0,
                p.avg_daily_hours or 0,
                p.day_hours or 0, p.night_hours or 0,
                p.weekly_allowance_hours or 0, p.annual_leave_hours or 0,
                p.holiday_hours or 0, p.labor_day_hours or 0,
                p.day_wage or 0, p.night_wage or 0,
                p.weekly_allowance_pay or 0, p.annual_leave_pay or 0,
                p.holiday_pay or 0, p.labor_day_pay or 0,
                p.gross_pay or 0,
                p.insurance_health or 0, p.insurance_care or 0,
                p.insurance_employment or 0, p.insurance_pension or 0,
                p.total_deduction or 0, p.net_pay or 0,
            ])

        # 열 너비
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# ──────────────────────────────────────────────────────
# 유틸 함수
# ──────────────────────────────────────────────────────

def _insurance_rate_year(year: int, month: int) -> int:
    """1~6월: 전년도 요율, 7~12월: 해당 연도 요율"""
    return year if month >= 7 else year - 1


def _get_insurance_rate(db: Session, rate_year: int) -> Optional[InsuranceRate]:
    return db.query(InsuranceRate).filter(InsuranceRate.year == rate_year).first()


def _calc_insurance(
    gross_pay: int,
    stored_health: int,
    stored_care: int,
    stored_employment: int,
    stored_pension: int,
    rate: Optional[InsuranceRate],
) -> tuple[Decimal, Decimal, Decimal, Decimal]:
    """
    공제 계산 — 저장된 값이 0이면 요율로 자동 계산
    (관리자가 직접 수정한 경우 저장값 우선)
    """
    gp = Decimal(gross_pay)

    health = Decimal(stored_health or 0)
    care = Decimal(stored_care or 0)
    employment = Decimal(stored_employment or 0)
    pension = Decimal(stored_pension or 0)

    if rate:
        if health == 0:
            health = (
                gp * rate.health_insurance_rate / Decimal("100")
            ).quantize(Decimal("1E1"), rounding=ROUND_DOWN)

        if care == 0:
            care = (
                health * rate.long_term_care_rate / Decimal("100")
            ).quantize(Decimal("1E1"), rounding=ROUND_DOWN)

        if employment == 0:
            employment = (
                gp * rate.employment_insurance_rate / Decimal("100")
            ).quantize(Decimal("1E1"), rounding=ROUND_DOWN)

        if pension == 0:
            pension = (
                gp * rate.national_pension_rate / Decimal("100")
            ).quantize(Decimal("1E1"), rounding=ROUND_DOWN)

    return health, care, employment, pension


def _count_work_days(db: Session, user_id: int, year: int, month: int) -> int:
    """해당 월 CLOCK_OUT 이벤트 기준 근무일 수"""
    month_start = date(year, month, 1)
    month_end = (
        date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)
    )
    return (
        db.query(AttendanceEvent.work_date)
        .filter(
            AttendanceEvent.user_id == user_id,
            AttendanceEvent.event_type == EventType.CLOCK_OUT,
            AttendanceEvent.work_date >= month_start,
            AttendanceEvent.work_date < month_end,
        )
        .distinct()
        .count()
    )


def _get_pay_date(db: Session, year: int, month: int) -> Optional[date]:
    pay_date = (
        db.query(PayrollPayDate)
        .filter(PayrollPayDate.year == year, PayrollPayDate.month == month)
        .first()
    )
    return pay_date.pay_date if pay_date else None
