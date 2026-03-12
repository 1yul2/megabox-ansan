"""
근태 라우터
─────────────────────────────────────────
1. 레거시 (username + password)   : POST /check-in, /break-start, /break-end, /check-out
2. 키오스크 (system JWT + user_id): GET /employees, GET /today/{user_id},
                                    POST /kiosk/check-in, /kiosk/break-start,
                                         /kiosk/break-end, /kiosk/check-out
3. 관리자                         : GET /admin/monthly, GET /admin/template,
                                    POST /admin/bulk-import
"""

from __future__ import annotations

import io
from datetime import date, datetime, time
from typing import List, Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_admin, get_current_user
from app.modules.auth.models import PositionEnum, StatusEnum, User
from app.modules.auth.services import verify_password
from app.modules.workstatus import models, schemas
from app.modules.workstatus.models import AttendanceEvent, EventType
from app.modules.workstatus.services import AttendanceService
from app.utils.permission_utils import is_system

router = APIRouter()

_KIOSK_POSITIONS = {PositionEnum.crew, PositionEnum.leader, PositionEnum.cleaner}


# ════════════════════════════════════════════════════════
# 공통 유틸
# ════════════════════════════════════════════════════════

def require_system_user(user: User = Depends(get_current_user)) -> User:
    if not is_system(user):
        raise HTTPException(status_code=403, detail="시스템 계정만 접근 가능합니다.")
    return user


def authenticate_attendance_user(db: Session, username: str, password: str) -> User:
    user = db.execute(
        select(User).where(User.username == username)
    ).scalar_one_or_none()

    if not user or not verify_password(password, user.password):
        raise HTTPException(status_code=401, detail="아이디 또는 비밀번호가 올바르지 않습니다.")
    if not user.is_active:
        raise HTTPException(status_code=400, detail="비활성화된 계정입니다.")
    return user


def _get_kiosk_user(db: Session, user_id: int) -> User:
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="직원을 찾을 수 없습니다.")
    if user.status != StatusEnum.approved:
        raise HTTPException(status_code=400, detail="승인된 계정이 아닙니다.")
    if user.position not in _KIOSK_POSITIONS:
        raise HTTPException(status_code=400, detail="근태 대상 직급이 아닙니다.")
    return user


def _add_event(
    db: Session,
    user_id: int,
    work_date: date,
    event_type: EventType,
    event_time: Optional[time] = None,
) -> AttendanceEvent:
    """이벤트 추가 (중복 시 409)"""
    if event_time is None:
        event_time = datetime.now().time()

    ev = AttendanceEvent(
        user_id=user_id,
        work_date=work_date,
        event_type=event_type,
        event_time=event_time,
    )
    db.add(ev)
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail=f"이미 {event_type.value} 기록이 있습니다.",
        )
    return ev


def _build_summary(db: Session, user_id: int, work_date: date) -> schemas.DailySummary:
    summary = AttendanceService.get_today_summary(db, user_id, work_date)
    user = db.get(User, user_id)
    if summary is None:
        return schemas.DailySummary(
            user_id=user_id,
            user_name=user.name if user else None,
            work_date=work_date,
        )
    events = AttendanceService.get_events_for_date(db, user_id, work_date)
    dm, nm, _ = AttendanceService.calc_work_minutes(events, work_date)
    return schemas.DailySummary(
        user_id=user_id,
        user_name=user.name if user else None,
        position=user.position.value if user else None,
        work_date=work_date,
        check_in=summary["check_in"],
        break_start=summary["break_start"],
        break_end=summary["break_end"],
        check_out=summary["check_out"],
        total_work_hours=float(
            AttendanceService.minutes_to_hours(dm + nm)
        ),
        day_hours=float(AttendanceService.minutes_to_hours(dm)),
        night_hours=float(AttendanceService.minutes_to_hours(nm)),
    )


# ════════════════════════════════════════════════════════
# 레거시 API (username + password)
# ════════════════════════════════════════════════════════

class AttendanceAuthInput(BaseModel):
    username: str
    password: str


@router.post("/check-in", response_model=schemas.DailySummary, summary="[레거시] 출근")
def check_in(
    payload: AttendanceAuthInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = authenticate_attendance_user(db, payload.username, payload.password)
    today = datetime.now().date()

    _add_event(db, user.id, today, EventType.CLOCK_IN)
    db.commit()
    return _build_summary(db, user.id, today)


@router.post("/break-start", response_model=schemas.DailySummary, summary="[레거시] 휴식 시작")
def break_start(
    payload: AttendanceAuthInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = authenticate_attendance_user(db, payload.username, payload.password)
    today = datetime.now().date()

    events = AttendanceService.get_events_for_date(db, user.id, today)
    if EventType.CLOCK_IN not in events:
        raise HTTPException(status_code=400, detail="출근 먼저 해주세요.")
    if EventType.CLOCK_OUT in events:
        raise HTTPException(status_code=400, detail="이미 퇴근한 기록이 있습니다.")

    _add_event(db, user.id, today, EventType.BREAK_START)
    db.commit()
    return _build_summary(db, user.id, today)


@router.post("/break-end", response_model=schemas.DailySummary, summary="[레거시] 복귀")
def break_end(
    payload: AttendanceAuthInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = authenticate_attendance_user(db, payload.username, payload.password)
    today = datetime.now().date()

    events = AttendanceService.get_events_for_date(db, user.id, today)
    if EventType.BREAK_START not in events:
        raise HTTPException(status_code=400, detail="휴식 시작 기록이 없습니다.")

    _add_event(db, user.id, today, EventType.BREAK_END)
    db.commit()
    return _build_summary(db, user.id, today)


@router.post("/check-out", response_model=schemas.DailySummary, summary="[레거시] 퇴근")
def check_out(
    payload: AttendanceAuthInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = authenticate_attendance_user(db, payload.username, payload.password)
    today = datetime.now().date()

    events = AttendanceService.get_events_for_date(db, user.id, today)
    if EventType.CLOCK_IN not in events:
        raise HTTPException(status_code=400, detail="출근 기록이 없습니다.")
    if EventType.BREAK_START in events and EventType.BREAK_END not in events:
        raise HTTPException(status_code=400, detail="휴식 중에는 퇴근할 수 없습니다.")

    _add_event(db, user.id, today, EventType.CLOCK_OUT)
    AttendanceService.handle_check_out(db, user.id, today)
    db.commit()
    return _build_summary(db, user.id, today)


# ════════════════════════════════════════════════════════
# 키오스크 API (system JWT + user_id)
# ════════════════════════════════════════════════════════

@router.get(
    "/employees",
    response_model=schemas.KioskEmployeesResponse,
    summary="[키오스크] 직원 목록",
)
def get_kiosk_employees(
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    users = (
        db.query(User)
        .filter(
            User.status == StatusEnum.approved,
            User.position.in_(_KIOSK_POSITIONS),
        )
        .order_by(User.name)
        .all()
    )
    return schemas.KioskEmployeesResponse(
        items=[
            schemas.KioskEmployeeDTO(
                id=u.id,
                name=u.name,
                position=u.position.value,
                username=u.username,
            )
            for u in users
        ]
    )


@router.get(
    "/today/{user_id}",
    response_model=Optional[schemas.DailySummary],
    summary="[키오스크] 오늘 근태 조회",
)
def get_today_kiosk(
    user_id: int,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    today = datetime.now().date()
    summary = AttendanceService.get_today_summary(db, user_id, today)
    if not summary:
        return None
    return _build_summary(db, user_id, today)


@router.post(
    "/kiosk/check-in",
    response_model=schemas.DailySummary,
    summary="[키오스크] 출근",
)
def kiosk_check_in(
    payload: schemas.KioskActionInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = _get_kiosk_user(db, payload.user_id)
    today = datetime.now().date()

    _add_event(db, user.id, today, EventType.CLOCK_IN)
    db.commit()
    return _build_summary(db, user.id, today)


@router.post(
    "/kiosk/break-start",
    response_model=schemas.DailySummary,
    summary="[키오스크] 휴식 시작",
)
def kiosk_break_start(
    payload: schemas.KioskActionInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = _get_kiosk_user(db, payload.user_id)
    today = datetime.now().date()

    events = AttendanceService.get_events_for_date(db, user.id, today)
    if EventType.CLOCK_IN not in events:
        raise HTTPException(status_code=400, detail="출근 먼저 해주세요.")
    if EventType.CLOCK_OUT in events:
        raise HTTPException(status_code=400, detail="이미 퇴근한 기록이 있습니다.")
    if EventType.BREAK_START in events and EventType.BREAK_END not in events:
        raise HTTPException(status_code=400, detail="이미 휴식 중입니다.")

    _add_event(db, user.id, today, EventType.BREAK_START)
    db.commit()
    return _build_summary(db, user.id, today)


@router.post(
    "/kiosk/break-end",
    response_model=schemas.DailySummary,
    summary="[키오스크] 복귀",
)
def kiosk_break_end(
    payload: schemas.KioskActionInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = _get_kiosk_user(db, payload.user_id)
    today = datetime.now().date()

    events = AttendanceService.get_events_for_date(db, user.id, today)
    if EventType.BREAK_START not in events:
        raise HTTPException(status_code=400, detail="휴식 시작 기록이 없습니다.")
    if EventType.BREAK_END in events:
        raise HTTPException(status_code=400, detail="이미 복귀한 기록이 있습니다.")

    _add_event(db, user.id, today, EventType.BREAK_END)
    db.commit()
    return _build_summary(db, user.id, today)


@router.post(
    "/kiosk/check-out",
    response_model=schemas.DailySummary,
    summary="[키오스크] 퇴근",
)
def kiosk_check_out(
    payload: schemas.KioskActionInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = _get_kiosk_user(db, payload.user_id)
    today = datetime.now().date()

    events = AttendanceService.get_events_for_date(db, user.id, today)
    if EventType.CLOCK_IN not in events:
        raise HTTPException(status_code=400, detail="출근 기록이 없습니다.")
    if EventType.BREAK_START in events and EventType.BREAK_END not in events:
        raise HTTPException(status_code=400, detail="휴식 중에는 퇴근할 수 없습니다.")
    if EventType.CLOCK_OUT in events:
        raise HTTPException(status_code=400, detail="이미 퇴근 기록이 있습니다.")

    _add_event(db, user.id, today, EventType.CLOCK_OUT)
    AttendanceService.handle_check_out(db, user.id, today)
    db.commit()
    return _build_summary(db, user.id, today)


# ════════════════════════════════════════════════════════
# 관리자 근태 관리 API
# ════════════════════════════════════════════════════════

@router.get(
    "/admin/monthly",
    response_model=schemas.MonthlyAttendanceResponse,
    summary="[관리자] 월별 근태 조회",
)
def admin_monthly_attendance(
    year: int = Query(...),
    month: int = Query(...),
    user_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    records = AttendanceService.get_monthly_attendance(db, year, month, user_id)
    return schemas.MonthlyAttendanceResponse(
        records=[schemas.DailySummary(**r) for r in records],
        total=len(records),
    )


@router.get(
    "/admin/template",
    summary="[관리자] 근태 엑셀 양식 다운로드",
)
def download_attendance_template(
    _admin=Depends(get_current_admin),
):
    """근태 대량 업로드용 엑셀 양식 다운로드"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "근태등록"

    headers = [
        "user_id",
        "name",
        "date",
        "clock_in",
        "break_start",
        "break_end",
        "clock_out",
    ]
    ws.append(headers)

    # 예시 데이터
    ws.append([1, "홍길동", "2026-03-01", "09:00", "12:00", "13:00", "18:00"])

    # 열 너비 조정
    for col_idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance_template.xlsx"},
    )


@router.post(
    "/admin/bulk-import",
    response_model=schemas.BulkImportResult,
    summary="[관리자] 근태 엑셀 대량 등록",
)
def bulk_import_attendance(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """
    엑셀 파일 업로드 → 근태 이벤트 자동 등록
    컬럼: user_id, name, date, clock_in, break_start, break_end, clock_out
    """
    try:
        content = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀 파일 파싱 오류: {e}")

    success_count = 0
    error_count = 0
    errors: List[str] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        try:
            user_id_raw, name_raw, date_raw, clock_in_raw, break_start_raw, break_end_raw, clock_out_raw = row[:7]

            if user_id_raw is None or date_raw is None or clock_in_raw is None or clock_out_raw is None:
                raise ValueError("필수 컬럼(user_id, date, clock_in, clock_out) 누락")

            user_id = int(user_id_raw)

            # 날짜 파싱
            if isinstance(date_raw, str):
                work_date = date.fromisoformat(date_raw.strip())
            else:
                work_date = date_raw  # Excel에서 date 객체로 읽히는 경우

            # 시간 파싱 (HH:MM 또는 HH:MM:SS 문자열)
            def parse_time(val) -> Optional[time]:
                if val is None:
                    return None
                if isinstance(val, time):
                    return val
                s = str(val).strip()
                for fmt in ("%H:%M:%S", "%H:%M"):
                    try:
                        return datetime.strptime(s, fmt).time()
                    except ValueError:
                        pass
                raise ValueError(f"시간 형식 오류: {val}")

            clock_in_time = parse_time(clock_in_raw)
            break_start_time = parse_time(break_start_raw)
            break_end_time = parse_time(break_end_raw)
            clock_out_time = parse_time(clock_out_raw)

            # 직원 존재 확인
            user = db.get(User, user_id)
            if not user:
                raise ValueError(f"user_id={user_id} 직원을 찾을 수 없습니다")

            # 기존 이벤트 삭제 (덮어쓰기)
            db.query(AttendanceEvent).filter_by(
                user_id=user_id, work_date=work_date
            ).delete()

            # 이벤트 등록
            if clock_in_time:
                db.add(AttendanceEvent(
                    user_id=user_id, work_date=work_date,
                    event_type=EventType.CLOCK_IN, event_time=clock_in_time,
                ))
            if break_start_time:
                db.add(AttendanceEvent(
                    user_id=user_id, work_date=work_date,
                    event_type=EventType.BREAK_START, event_time=break_start_time,
                ))
            if break_end_time:
                db.add(AttendanceEvent(
                    user_id=user_id, work_date=work_date,
                    event_type=EventType.BREAK_END, event_time=break_end_time,
                ))
            if clock_out_time:
                db.add(AttendanceEvent(
                    user_id=user_id, work_date=work_date,
                    event_type=EventType.CLOCK_OUT, event_time=clock_out_time,
                ))

            db.flush()

            # Payroll 처리
            if clock_in_time and clock_out_time:
                AttendanceService.handle_check_out(db, user_id, work_date)

            success_count += 1

        except Exception as e:
            error_count += 1
            errors.append(f"Row {row_idx}: {e}")
            db.rollback()
            # 오류 행만 건너뜀, 계속 진행

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"DB 저장 오류: {e}")

    return schemas.BulkImportResult(
        success_count=success_count,
        error_count=error_count,
        errors=errors,
    )
